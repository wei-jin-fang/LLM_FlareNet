import argparse
import sys

import matplotlib
import torch

import random
import numpy as np
import os

import time

import openpyxl
import torch
import torch.optim as optim
import torch.nn.functional as F
import numpy as np
import argparse
import os
import pandas as pd
from sklearn.utils import compute_class_weight
from tools import BS_BSS_score, BSS_eval_np

from tools import Metric, plot_losses
from tools import getClass
from tools import shuffle_data
from tools import get_batches_integer_test
from tools import Rectify_binary
from tools import save_torchModel
from tools import setup_seed_torch
from tools import DualOutput
from tools import save_csv

os.environ['CURL_CA_BUNDLE'] = ''
os.environ["PYTORCH_CUDA_ALLOC_CONF"] = "max_split_size_mb:64"


def load_data(train_csv_path, validate_csv_path, test_csv_path, Class_NUM):
    pd.set_option('display.max_columns', None)  # 意思是不限制显示的列数。这样设置后，无论 Pandas 数据帧有多少列
    List = []
    count_index = 0
    for path in [train_csv_path, validate_csv_path, test_csv_path]:
        count_index += 1
        csv = pd.read_csv(path)
        start = 0
        end = 0
        for Column in csv.columns.values:
            start += 1
            if Column.__eq__("TOTUSJH"):
                break

        for Column in csv.columns.values:
            end += 1
            if Column.__eq__("SHRGT45"):
                break
        List.append(csv.iloc[:, start - 1:end].values)  # train_x  test_x
        Classes = csv["CLASS"].copy()
        # 定义分类数组
        categories = ["NC", "MX"]
        # 初始化类别计数列表
        weights = [0] * len(categories)
        class_list = []

        # 遍历每个Class，计算每个类别的数量
        for Class_ in Classes:
            for i, category in enumerate(categories):
                if Class_ in category:
                    weights[i] += 1
                    class_list.append(i)
                    break  # 找到匹配的类别后退出内层循环
        class_tensor = torch.tensor(class_list, dtype=torch.long)
        List.append(np.array(class_tensor))

        print(f"{path}每一类的数量是:", weights)
        tempweight = []

        for weight in weights:
            tempweight.append(weight)

        weight_list = getClass(tempweight)
        print(f"{path}get_Class函数得到的的权重：", weight_list)
        # exit()
    return List[0], List[1], List[2], List[3], List[4], List[5]


def Preprocess(train_csv_path, validate_csv_path, test_csv_path):
    global FIRST

    train_x, train_y, validate_x, validate_y, test_x, test_y = load_data(train_csv_path, validate_csv_path,
                                                                         test_csv_path, Class_NUM)
    # print(train_x.shape)(17800, 10)
    train_x = train_x.reshape(-1, TIME_STEPS, INPUT_SIZE)
    # print(train_x.shape)(445, 40, 10)

    train_y = Rectify_binary(train_y, Class_NUM, TIME_STEPS)
    validate_x = validate_x.reshape(-1, TIME_STEPS, INPUT_SIZE)
    test_x = test_x.reshape(-1, TIME_STEPS, INPUT_SIZE)
    validate_y = Rectify_binary(validate_y, Class_NUM, TIME_STEPS)
    test_y = Rectify_binary(test_y, Class_NUM, TIME_STEPS)

    if FIRST == 1:
        print("train_x.shape : {} ".format(train_x.shape))
        print("train_y.shape : {} ".format(train_y.shape))
        print("validate_x.shape : {} ".format(validate_x.shape))
        print("validate_y.shape : {} ".format(validate_y.shape))
        print("test_x.shape : {} ".format(test_x.shape))
        print("test_y.shape : {} ".format(test_y.shape))
        FIRST = 0

    def class_weight(alpha, beta, zero=True):
        # 权重计算-修改位置开始
        classes = np.array([0, 1])
        # 权重计算-修改位置结束
        # weight = compute_class_weight(class_weight='balanced', classes=classes, y=train_y.argmax(axis=1))
        # exit()
        weight = compute_class_weight(class_weight='balanced', classes=classes, y=train_y)
        if zero:
            weight[0] = weight[0] * alpha
            print("zero:", weight[0])
            return weight[0]
        else:
            weight[1] = weight[1] * beta
            print("zero:", weight[1])
            return weight[1]

    class_weight = {0.: class_weight(alpha=1, beta=1, zero=True), 1.: class_weight(alpha=1, beta=1, zero=False)}
    print(class_weight)

    return train_x, train_y, validate_x, validate_y, test_x, test_y, class_weight


def train(ep, train_x, train_y, optimizer, model):
    global steps
    train_loss = 0
    model.train()
    batch_count = 0
    for batch_idx, (data, target) in enumerate(get_batches_integer_test(train_x, train_y, batch_size)):
        # exit()
        if not isinstance(data, torch.Tensor):
            data = torch.tensor(data, dtype=torch.float32)
        if not isinstance(target, torch.Tensor):
            target = torch.tensor(target, dtype=torch.long)
        # data = data.unsqueeze(1)  # 增加维度适应
        if args.cuda:
            data, target = data.cuda(), target.cuda()
        # data = data.contiguous().view(-1, 1, INPUT_SIZE * TIME_STEPS)
        if args.cuda:
            data, target = data.cuda(), target.cuda()
            class_weights_cuda = torch.tensor([class_weight[0.], class_weight[1.]], dtype=torch.float32).cuda()
        else:
            class_weights_cuda = torch.tensor([class_weight[0.], class_weight[1.]], dtype=torch.float32)

        optimizer.zero_grad()
        # print("data.shape",data.shape) #data.shape torch.Size([4, 1, 400])
        output = model(data)
        '''
                loss = F.binary_cross_entropy_with_logits(output, target, weight=class_weights_cuda)
        '''
        # exit()
        loss = F.nll_loss(output, target, weight=class_weights_cuda)
        loss.backward()
        # if args.clip > 0:
        #     torch.nn.utils.clip_grad_norm_(models.parameters(), args.clip)
        optimizer.step()
        train_loss += loss.item()
        steps += len(data)
        message = ('Train Epoch: {} \tLoss: {:.6f}\tSteps: {}'.format(
            ep, train_loss / args.log_interval, steps))
        batch_count += 1
    return train_loss / batch_count


def evaluate(data_x, data_y, model):
    model.eval()
    test_loss = 0
    correct = 0
    all_targets = []
    all_predictions = []
    batch_count = 0

    all_predictions_y_true = []
    all_predictions_y_prob = []
    with torch.no_grad():
        for data, target in get_batches_integer_test(data_x, data_y, batch_size):

            if not isinstance(data, torch.Tensor):
                data = torch.tensor(data, dtype=torch.float32)
            if not isinstance(target, torch.Tensor):
                target = torch.tensor(target, dtype=torch.long)

            # data = data.unsqueeze(1)  # 增加维度适应
            if args.cuda:
                data, target = data.cuda(), target.cuda()
            # data = data.contiguous().view(-1, 1, INPUT_SIZE * TIME_STEPS)  # 调整为适合模型输入的形状
            if args.cuda:
                data, target = data.cuda(), target.cuda()
                class_weights_cuda = torch.tensor([class_weight[0.], class_weight[1.]], dtype=torch.float32).cuda()
            else:
                class_weights_cuda = torch.tensor([class_weight[0.], class_weight[1.]], dtype=torch.float32)
            output = model(data)
            # test_loss += F.nll_loss(output, target, weight=class_weights_cuda, size_average=False).item()
            test_loss += F.nll_loss(output, target, size_average=False).item()

            pred = torch.argmax(output, dim=1)  # 拿到概率最大的下标
            correct += pred.eq(target.data.view_as(pred)).cpu().sum()
            all_targets.extend(target.cpu().numpy())  # 这一批次实际label加进去数组
            all_predictions.extend(torch.argmax(output, dim=1).cpu().numpy())  # 这一批次预测label加进去数组

            # 添加内容方便计算BSS BS
            probabilities = torch.exp(output)  # 拿到这一批次概率数值
            # tensor([[0.8139, 0.1861],
            #         [0.8321, 0.1679],
            # all_predictions_y_true.extend(torch.argmax(output, dim=1).cpu().numpy())  # 拿到预测的实际label
            all_predictions_y_true.extend(target.cpu().numpy())  # 拿到预测的实际label
            all_predictions_y_prob.extend(probabilities.cpu().numpy().tolist())  # 拿到概率数值加入数组

            batch_count += 1
        metric = Metric(all_targets, all_predictions)
        print(metric.Matrix())
        Tss = metric.TSS()[0]
        print(f"TSS: {Tss}")
        test_loss /= batch_count
        accuracy = correct / len(data_x)
        message = ('\nTest set: Average loss: {:.4f}, Accuracy: {}/{} ({:.0f}%)\n'.format(
            test_loss, correct, len(data_x),
            100. * accuracy))
        print("本轮评估完成", message)
        return {'loss': test_loss, 'accuracy': accuracy, 'tss': Tss,
                "metric": metric}, all_predictions_y_true, all_predictions_y_prob


import shap
import matplotlib.pyplot as plt

matplotlib.use('Agg')
import numpy as np
import torch
import shap
import matplotlib.pyplot as plt


def SHAP_all(train_x,data_x, data_y, model, count, model_typ):
    # 确保输入数据为 NumPy 数组

    if isinstance(data_x, torch.Tensor):
        data_x = data_x.cpu().numpy()
    if isinstance(train_x, torch.Tensor):
        train_x = train_x.cpu().numpy()

    # 转换为 float32 类型
    data_x = data_x.astype(np.float32)

    # 选择背景数据（若数据量大，可用子集以减少计算量）
    background_data = train_x
    background_tensor = torch.tensor(background_data, dtype=torch.float32)

    # 如果有 CUDA，将背景数据和模型移到 GPU
    args = type('Args', (), {'cuda': torch.cuda.is_available()})()  # 模拟 args.cuda
    if args.cuda:
        background_tensor = background_tensor.cuda()
        model = model.cuda()

    # 手动计算 expected_value（正类的平均模型输出）
    model.eval()  # 设置模型为评估模式
    with torch.no_grad():
        # base value统一用模型的输出
        background_outputs = model(background_tensor).cpu().numpy()  # 转换为概率，形状: (n_background, n_classes)
        # base value统一用模型的输出
        expected_value = np.mean(background_outputs, axis=0)  # 形状: (n_classes,)
        classnumber = 1
        expected_value_positive = expected_value[classnumber]  # 正类的标量值
    print(f"正类的预期值: {expected_value_positive}")

    # 创建特征名称
    feature_names = ["TOTUSJH", "TOTPOT", "TOTUSJZ", "ABSNJZH", "SAVNCPP",
                     "USFLUX", "AREA_ACR", "MEANPOT", "R_VALUE", "SHRGT45"]

    # 初始化 SHAP 解释器
    explainer = shap.GradientExplainer(model, background_tensor)

    # 初始化存储所有样本的 SHAP 值
    n_samples = data_x.shape[0]
    shap_values_positive_mean_all = []

    # 循环处理每个样本
    for noaaid_number in range(n_samples):
        # 提取单个样本数据
        single_sample = data_x[noaaid_number:noaaid_number+1]  # 形状: (1, 40, 10)
        single_tensor = torch.tensor(single_sample, dtype=torch.float32)
        if args.cuda:
            single_tensor = single_tensor.cuda()

        # 计算单个样本的模型输出并应用指数操作
        with torch.no_grad():
            nowprob = model(single_tensor).cpu().numpy()  # 形状: (1, n_classes)
            # 概率统一用模型输出
            # nowprob_exp = np.exp(nowprob)  # 指数操作
            # nowprob_positive_exp =nowprob_exp
        print(f"样本 {noaaid_number} 的模型输出 (nowprob): {nowprob}")

        # 计算单个样本的 SHAP 值
        print(f"Computing SHAP values for sample {noaaid_number}...")
        shap_values = explainer.shap_values(single_tensor)
        print(f"SHAP values shape: {np.array(shap_values).shape}")  # 形状: (2, 1, 40, 10)

        # 提取正类的 SHAP 值
        shap_values_positive = shap_values[classnumber]  # 形状: (1, 40, 10)
        # 提取最后一个时间步长的 SHAP 值
        # shap_values_positive_mean = shap_values_positive[:, -1, :]  # 形状: (1, 10)
        # 按时间步（40）取均值，得到每个特征的 SHAP 值
        shap_values_positive_mean = np.mean(shap_values_positive, axis=1)  # 形状: (1, 10)
        shap_values_positive_mean = pd.DataFrame(shap_values_positive_mean, columns=feature_names)



        # 输入数据按时间步取均值
        data_x_mean = np.mean(single_sample, axis=1)  # 形状: (1, 10)
        # 提取最后一个时间步长的 SHAP 值
        # data_x_mean = single_sample[:, -1, :]  # 形状: (1, 10)
        data_x_mean = pd.DataFrame(data_x_mean, columns=feature_names)

        # 存储 SHAP 值
        shap_values_positive_mean_all.append(shap_values_positive_mean)

        # 创建 SHAP Explanation 对象用于瀑布图
        '''
        shap.Explanation 的 values 参数期望一个一维数组，表示单个样本的 SHAP 值。
        由于 shap_values_positive_mean 是一个 DataFrame（即使只有一行），直接使用 shap_values_positive_mean 会包含 DataFrame 的结构（行和列），而不是单纯的数值数组。
        iloc[0] 提取第 0 行的数据（因为你只处理一个样本，noaaid_number:noaaid_number+1 确保每次循环只有一个样本），然后 .values 提取数值部分
        '''
        explanation = shap.Explanation(
            values=shap_values_positive_mean.iloc[0].values,  # SHAP 值 (10,)
            base_values=expected_value_positive,  # 基线值
            data=data_x_mean.iloc[0],  # 特征值 (10,)
            feature_names=feature_names

        )

        # 绘制 SHAP 瀑布图
        plot_type = "waterfall不是最新"
        shap.waterfall_plot(explanation, show=False)  # show=False 以便保存

        # 设置标题
        plt.title(f"SHAP Waterfall Plot for Positive Class (Sample {noaaid_number}, Model: {model_typ})")

        # 保存图像
        filename = f"tesing/waterfall不是最新/{plot_type}-{model_typ}-positive_class-{classnumber}-summary_plot-{count}({noaaid_number})-({nowprob}).png"
        plt.savefig(filename, bbox_inches="tight", dpi=300)
        print(f"图像已保存至 {filename}")

        # 释放内存
        plt.close()

    # 合并所有样本的 SHAP 值
    shap_values_positive_mean_all = pd.concat(shap_values_positive_mean_all, ignore_index=True)



parser = argparse.ArgumentParser(description='Time-LLM')

# basic config

parser.add_argument('--embed', type=str, default='timeF',
                    help='time features encoding, options:[timeF, fixed, learned]')

# my参数
parser.add_argument('--dropout', type=float, default=0.5, help='dropout')
parser.add_argument('--cuda', action='store_false')
parser.add_argument('--epochs', type=int, default=160)
parser.add_argument('--num_dataset', type=int, default=9)  # 循环处理0-9个数据集
parser.add_argument('--batch_size', type=int, default=16, help='batch size of train input data')
parser.add_argument('--num_hidden_layers', type=int, default=4)
parser.add_argument('--num_heads', type=int, default=5)  # 大模型层
parser.add_argument('--log-interval', type=int, default=100, metavar='N')
parser.add_argument('--lr', type=float, default=0.001)
parser.add_argument('--optim', type=str, default='Adam')
parser.add_argument('--seed', type=int, default=2021, help='random seed')
parser.add_argument('--model_type', type=str, default='VIT', help='VIT,LLM_VIT')
parser.add_argument('--embed_dim', type=str, default=100, help='VIT,LLM_VIT')  # 大模型层
parser.add_argument('--hidden_units', type=int, default=256, help='64')
parser.add_argument('--num_layers', type=int, default=2, help='64')
parser.add_argument('--print', type=int, default=1, help='是否第一次输出')
parser.add_argument('--datasetname', type=str, default="data", help='new_data_scaler,data')
args = parser.parse_args()

device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
batch_size = args.batch_size

FIRST = 1
# 参数设置-修改位置开始
TIME_STEPS = 40
INPUT_SIZE = 10
Class_NUM = 2

# 参数设置-修改位置结束

import LLM_VIT

if __name__ == "__main__":
    # for item in range(8,120,8):
    #     args.embed_dim=item
    #     for _modelType in ["VIT","LSTM_new","LLM","LSTM"]:
    for _modelType in ["LLM_VIT"]:
        # for _modelType in ["LLM_VIT"]:
        # for _modelType in ["LLM_VIT"]:
        args.model_type = _modelType
        other = "全局归一化数据集_监测TSS"
        if args.datasetname == "new_data_scaler":
            other = "第三种归一化数据集_监测TSS"
        commment = 'lr{}_epochs{}_datasetNum{}_modelType{}_batchsize{}_numhiddenlayers{}_lr{}_numheads{}_embeddim{}_hiddenunits{}_numlayers{}_datasetname{}'.format(
            args.lr,
            args.epochs,
            args.num_dataset,
            args.model_type,
            args.batch_size,
            args.num_hidden_layers,
            args.lr,
            args.num_heads,
            args.embed_dim,
            args.hidden_units,
            args.num_layers,
            args.datasetname,
            other
        )
        print(commment)
        start_time = time.time()

        timelabel = time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime(time.time()))
        model_base = f"./model_output/{timelabel}"
        os.makedirs(f"{model_base}")
        os.makedirs(f"{model_base}/plot")
        results_filepath = f"{model_base}/important.txt"
        results_logfilepath = f"{model_base}/log.txt"

        sys.stdout = DualOutput(results_logfilepath)

        # 定义十个数据集合存储内容
        all_matrix = np.array([[0, 0], [0, 0]])
        data_Recall, data_Precision, data_Accuracy, data_TSS, data_BSS, data_HSS, data_FAR = [], [], [], [], [], [], []
        data_FPR = []
        # 打开文件准备写入
        with open(results_filepath, 'w') as results_file:
            # for count in range(10):  # 循环处理0-9个数据集
            # for count in range(args.num_dataset + 1):  # 循环处理0-9个数据集
            for count1 in range(args.num_dataset + 1):  # 循环处理0-9个数据集
                setup_seed_torch(args.seed)

                steps = 0
                print(args)
                count = 1
                # dataname="new_data_scaler"
                dataname = args.datasetname
                train_csv_path = rf"../../../{dataname}/{count}Train.csv"
                validate_csv_path = rf"../../../{dataname}/{count}Val.csv"
                test_csv_path = rf"../../../{dataname}/{count}Test.csv"
                # train_csv_path = rf"E:\conda_code_tf\LLM\LLM_VIT\Data_origin\1Test.csv"
                # validate_csv_path = rf"E:\conda_code_tf\LLM\LLM_VIT\Data_origin\1Test.csv"
                # test_csv_path = rf"E:\conda_code_tf\LLM\LLM_VIT\Data_origin\1Test.csv"
                train_x, train_y, validate_x, validate_y, test_x, test_y, class_weight = Preprocess(train_csv_path,
                                                                                                    validate_csv_path,
                                                                                                    test_csv_path)

                model_filename = rf"G:\本科\项目_比赛_论文资料\论文_LLM_VIT\Titan\TSS" \
                                 + f"\TSS_{args.model_type}" + f"\model_{count}.pt"

                # 加载最佳模型
                model_test = torch.load(model_filename)

                # 定义一个包装模型类
                class WrappedModel(torch.nn.Module):
                    def __init__(self, original_model):
                        super(WrappedModel, self).__init__()
                        self.model = original_model  # 加载的原始模型

                    def forward(self, x):
                        # 调用原始模型的前向传播
                        attention_mul = self.model(x)  # 假设原始模型输出 attention_mul
                        # 原始模型的输出是 log_softmax
                        log_probs = F.log_softmax(attention_mul, dim=1)
                        # 转换为 exp 后的概率值
                        probs = torch.exp(log_probs)
                        return probs


                # 包装模型
                wrapped_model = WrappedModel(model_test)

                model_test=wrapped_model



                # exit()
                # 测试集评估
                print(f"====================数据集{count}测试集轮评估数据=============================================")
                # test_metrics, all_predictions_y_true, all_predictions_y_prob = evaluate(test_x, test_y,
                #                                                                         model_test)  # 使用测试集进行评估

                SHAP_all(train_x,test_x, test_y, model_test, count, args.model_type)  # 使用测试集进行评估
                exit()
        #         continue
        #         del model_test
        #         # 清理 GPU 内存
        #         # torch.cuda.empty_cache()
        #
        #         # 保存csv
        #         y_true=np.array(all_predictions_y_true)#(160,)
        #         y_pred=np.array(all_predictions_y_prob)#(160, 2)
        #         path=fr"TSS_{args.model_type}_{count}"
        #         save_csv(y_pred,y_true,"./TSSCSV/",path)
        #         # exit()
        #
        #
        #         testMetrics = test_metrics["metric"]
        #         print(testMetrics.Matrix())
        #         all_matrix += testMetrics.Matrix()
        #
        #
        #         data_FPR.append(testMetrics.FPR())
        #         print("data_FPR", testMetrics.FPR())
        #
        #         data_Recall.append(testMetrics.Recall())
        #         print("Recall", testMetrics.Recall())
        #
        #         data_Precision.append(testMetrics.Precision())
        #         print("Precision", testMetrics.Precision())
        #
        #         data_Accuracy.append(testMetrics.Accuracy())
        #         print("Accuracy", testMetrics.Accuracy())
        #
        #         data_TSS.append(testMetrics.TSS())
        #         print("TSS", testMetrics.TSS())
        #
        #         data_HSS.append(testMetrics.HSS())
        #         print("HSS", testMetrics.HSS())
        #
        #         data_FAR.append(testMetrics.FAR())
        #         print("FAR", testMetrics.FAR())
        #         # 开始求BSS
        #         y_true = all_predictions_y_true
        #         y_prob = np.array([row[1] for row in all_predictions_y_prob])
        #
        #         BS, BSS = BS_BSS_score(y_true, y_prob)
        #         data_BSS.append([BS, BSS])
        #         print("BS, BSS", [BS, BSS])
        #
        #         print(f"数据集 {count} 测试集TSS:", test_metrics['tss'])
        #
        #         # 写入结果到文件
        #         results_file.write(f"数据集 {count} 测试集TSS: {test_metrics['tss']}\n")
        #         results_file.write("=================================================================\n")
        #
        #         # 绘制并保存损失曲线
        #
        #
        #     print("#接下来计算所有测试集指标均值和方法")
        #     # 转换数据为numpy数组以便计算
        #     data_Recall = np.array(data_Recall)
        #     data_Precision = np.array(data_Precision)
        #     data_Accuracy = np.array(data_Accuracy)
        #     data_TSS = np.array(data_TSS)
        #     data_HSS = np.array(data_HSS)
        #     data_FAR = np.array(data_FAR)
        #     data_BSS = np.array(data_BSS)
        #     data_FPR = np.array(data_FPR)
        #     # 计算均值和标准差
        #     results = {
        #         "Metric": ["Recall", "Precision", "Accuracy", "TSS", "HSS", "FAR", "BS/BSS","FPR"],
        #         "Mean": [data_Recall.mean(axis=0), data_Precision.mean(axis=0), data_Accuracy.mean(axis=0),
        #                  data_TSS.mean(axis=0), data_HSS.mean(axis=0), data_FAR.mean(axis=0), data_BSS.mean(axis=0),
        #                  data_FPR.mean(axis=0)],
        #         "Std": [data_Recall.std(axis=0), data_Precision.std(axis=0), data_Accuracy.std(axis=0),
        #                 data_TSS.std(axis=0), data_HSS.std(axis=0), data_FAR.std(axis=0), data_BSS.std(axis=0),
        #                 data_FPR.std(axis=0)]
        #     }
        #
        #     # 转换均值和标准差为"均值 ± 标准差"格式，并整理为DataFrame
        #     formatted_results = []
        #     for metric, means, stds in zip(results['Metric'], results['Mean'], results['Std']):
        #         neg_str = f"{str(means[0])[:str(means[0]).find('.') + 4]} ± {str(stds[0])[:str(stds[0]).find('.') + 4]}"
        #         pos_str = f"{str(means[1])[:str(means[1]).find('.') + 4]} ± {str(stds[1])[:str(stds[1]).find('.') + 4]}"
        #         formatted_results.append([metric, neg_str, pos_str])
        #
        #     # 创建DataFrame
        #     df = pd.DataFrame(formatted_results, columns=["Metric", "负类", "正类"])
        #     excel_filename = f'{model_base}/results.xlsx'
        #     df.to_excel(excel_filename, index=False)
        #     print(f"结果已写入 {excel_filename}")
        #
        #     # 记录当前TSS
        #
        #     filename = './result.xlsx'
        #     new_row = [f'{model_base}-{commment}', data_TSS.mean(axis=0)[1]]
        #
        #     # 判断文件是否存在
        #     if os.path.exists(filename):
        #         # 文件存在，加载文件
        #         wb = openpyxl.load_workbook(filename)
        #         ws = wb.active
        #     else:
        #         # 文件不存在，创建一个新工作簿
        #         wb = openpyxl.Workbook()
        #         ws = wb.active
        #
        #     # 获取当前行数
        #     current_row = ws.max_row + 1
        #
        #     # 将数据写入新行
        #     for col_num, value in enumerate(new_row, start=1):
        #         ws.cell(row=current_row, column=col_num, value=value)
        #
        #     # 保存文件
        #     wb.save(filename)
        #
        # # 删除模型
        # for filename in os.listdir(model_base):
        #     # 检查文件是否以 .pt 结尾
        #     if filename.endswith('.pt'):
        #         # 构造文件的完整路径
        #         file_path = os.path.join(model_base, filename)
        #         try:
        #             # 删除文件
        #             # os.remove(file_path)
        #             pass
        #         except Exception as e:
        #             print(f"删除文件 {file_path} 时出错: {e}")
        #
        # end_time = time.time()
        # elapsed_time_minutes = (end_time - start_time) / 60
        # print(f"程序运行时间: {elapsed_time_minutes:.2f} 分钟")
