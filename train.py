import argparse
import sys

import time
import openpyxl
import torch
import torch.optim as optim
import torch.nn.functional as F
import numpy as np
import argparse
import os
import pandas as pd
from sklearn.utils import compute_class_weight

from models.LLMFlareNet_1 import LLMFlareNet_1Model
from models.LLMFlareNet_2 import LLMFlareNet_2Model
from models.Onefitall import OnefitallModel
from models.Onefitall_11 import Onefitall_11Model
from models.Onefitall_12 import Onefitall_12Model
from models.Onefitall_13 import Onefitall_13Model
from tools import BS_BSS_score, BSS_eval_np, get_batches_all

from tools import Metric, plot_losses
from tools import getClass
from tools import shuffle_data
from tools import get_batches_integer

from tools import Rectify_binary
from tools import save_torchModel
from tools import setup_seed_torch
from tools import DualOutput

from models.LLMFlareNet import LLMFlareNetModel
'''
全局变量设置：
'''
TIME_STEPS = 40
INPUT_SIZE = 10
Class_NUM = 2
FIRST = 1


os.environ['CURL_CA_BUNDLE'] = ''
os.environ["PYTORCH_CUDA_ALLOC_CONF"] = "max_split_size_mb:64"

def load_data(train_csv_path, validate_csv_path, test_csv_path, Class_NUM):
    pd.set_option('display.max_columns', None)  # 意思是不限制显示的列数。这样设置后，无论 Pandas 数据帧有多少列
    List = []
    count_index = 0
    for path in [train_csv_path, validate_csv_path, test_csv_path]:
        count_index += 1
        csv = pd.read_csv(path)
        start = 0
        end = 0
        for Column in csv.columns.values:
            start += 1
            if Column.__eq__("TOTUSJH"):
                break

        for Column in csv.columns.values:
            end += 1
            if Column.__eq__("SHRGT45"):
                break
        List.append(csv.iloc[:, start - 1:end].values)  # train_x  test_x
        Classes = csv["CLASS"].copy()
        # 定义分类数组
        categories = ["NC", "MX"]
        # 初始化类别计数列表
        weights = [0] * len(categories)
        class_list = []

        # 遍历每个Class，计算每个类别的数量
        for Class_ in Classes:
            # 针对每一个数据
            for i, category in enumerate(categories):
                # 判断数据那个类别
                if Class_ in category:
                    weights[i] += 1
                    class_list.append(i)
                    break  # 找到匹配的类别后退出内层循环

        class_tensor = torch.tensor(class_list, dtype=torch.long)
        List.append(np.array(class_tensor))

        print(f"{path}每一类的数量是:", weights)
        tempweight = []

        for weight in weights:
            tempweight.append(weight)

        weight_list = getClass(tempweight)
        print(f"{path}get_Class函数得到的的权重：", weight_list) #用于测试

    return List[0], List[1], List[2], List[3], List[4], List[5]

def Preprocess(train_csv_path, validate_csv_path, test_csv_path):
    global FIRST

    train_x, train_y, validate_x, validate_y, test_x, test_y = \
        load_data(train_csv_path, validate_csv_path,test_csv_path, Class_NUM)

    train_x = train_x.reshape(-1, TIME_STEPS, INPUT_SIZE)
    train_y = Rectify_binary(train_y, Class_NUM, TIME_STEPS)
    validate_x = validate_x.reshape(-1, TIME_STEPS, INPUT_SIZE)
    test_x = test_x.reshape(-1, TIME_STEPS, INPUT_SIZE)
    validate_y = Rectify_binary(validate_y, Class_NUM, TIME_STEPS)
    test_y = Rectify_binary(test_y, Class_NUM, TIME_STEPS)

    if FIRST == 1:
        print("train_x.shape : {} ".format(train_x.shape))
        print("train_y.shape : {} ".format(train_y.shape))
        print("validate_x.shape : {} ".format(validate_x.shape))
        print("validate_y.shape : {} ".format(validate_y.shape))
        print("test_x.shape : {} ".format(test_x.shape))
        print("test_y.shape : {} ".format(test_y.shape))
        FIRST = 0

    def class_weight(alpha, beta, zero=True):
        # 权重计算-修改位置开始
        classes = np.array([0, 1])
        # 权重计算-修改位置结束
        weight = compute_class_weight(class_weight='balanced', classes=classes, y=train_y)
        if zero:
            weight[0] = weight[0] * alpha
            print("zero:", weight[0])
            return weight[0]
        else:
            weight[1] = weight[1] * beta
            print("zero:", weight[1])
            return weight[1]

    class_weight = {0.: class_weight(alpha=1, beta=1, zero=True), 1.: class_weight(alpha=1, beta=1, zero=False)}
    print(class_weight)

    return train_x, train_y, validate_x, validate_y, test_x, test_y, class_weight
def train_integer(ep, train_x, train_y, optimizer, model, batch_size):
    global steps
    train_loss = 0
    model.train()
    batch_count = 0
    for batch_idx, (data, target) in enumerate(get_batches_integer(train_x, train_y, batch_size)):
        if not isinstance(data, torch.Tensor):
            data = torch.tensor(data, dtype=torch.float32)
        # 修改 target 处理：无论是否为 Tensor，都强制转换为 float32
        target = torch.tensor(target, dtype=torch.float32) if not isinstance(target, torch.Tensor) else target.to(dtype=torch.float32)

        if args.cuda:
            data, target = data.cuda(), target.cuda()
            class_weights_cuda = torch.tensor([class_weight[1.]], dtype=torch.float32).cuda()  # 正类权重
        else:
            class_weights_cuda = torch.tensor([class_weight[1.]], dtype=torch.float32)

        optimizer.zero_grad()
        output = model(data)  # [batch_size, 1]，概率值
        target = target.view(-1, 1)  # 确保 target 形状为 [batch_size, 1]
        loss = F.binary_cross_entropy(output, target, weight=class_weights_cuda)  # 使用 BCE Loss
        loss.backward()
        optimizer.step()
        train_loss += loss.item()
        batch_count += 1

    message = ('Train Epoch: {} \t average Loss: {:.6f}'.format(ep, train_loss / batch_count))
    print(message)
    return train_loss / batch_count

def train_all(ep, train_x, train_y, optimizer, model,batch_size):
    global steps
    train_loss = 0
    model.train()
    batch_count = 0
    for batch_idx, (data, target) in enumerate(get_batches_all(train_x, train_y, batch_size)):
        if not isinstance(data, torch.Tensor):
            data = torch.tensor(data, dtype=torch.float32)
        # 修改 target 处理：无论是否为 Tensor，都强制转换为 float32
        target = torch.tensor(target, dtype=torch.float32) if not isinstance(target, torch.Tensor) else target.to(dtype=torch.float32)

        if args.cuda:
            data, target = data.cuda(), target.cuda()
            class_weights_cuda = torch.tensor([class_weight[1.]], dtype=torch.float32).cuda()  # 正类权重
        else:
            class_weights_cuda = torch.tensor([class_weight[1.]], dtype=torch.float32)

        optimizer.zero_grad()
        output = model(data)  # [batch_size, 1]，概率值
        target = target.view(-1, 1)  # 确保 target 形状为 [batch_size, 1]
        loss = F.binary_cross_entropy(output, target, weight=class_weights_cuda)  # 使用 BCE Loss
        loss.backward()
        optimizer.step()
        train_loss += loss.item()
        batch_count += 1

    message = ('Train Epoch: {} \t average Loss: {:.6f}'.format(ep, train_loss / batch_count))
    print(message)
    return train_loss / batch_count

def evaluate_integer(data_x, data_y, model, batch_size):
    model.eval()
    test_loss = 0
    correct = 0
    all_targets = []
    all_predictions = []
    batch_count = 0
    all_predictions_y_true = []
    all_predictions_y_prob = []

    with torch.no_grad():
        for data, target in get_batches_integer(data_x, data_y, batch_size):
            if not isinstance(data, torch.Tensor):
                data = torch.tensor(data, dtype=torch.float32)
            # 修改 target 处理：无论是否为 Tensor，都强制转换为 float32
            target = torch.tensor(target, dtype=torch.float32) if not isinstance(target, torch.Tensor) else target.to(dtype=torch.float32)

            if args.cuda:
                data, target = data.cuda(), target.cuda()
                class_weights_cuda = torch.tensor([class_weight[1.]], dtype=torch.float32).cuda()
            else:
                class_weights_cuda = torch.tensor([class_weight[1.]], dtype=torch.float32)

            output = model(data)  # [batch_size, 1]，概率值
            target = target.view(-1, 1)
            test_loss += F.binary_cross_entropy(output, target, weight=class_weights_cuda).item()

            pred = (output > 0.5).float()  # 阈值 0.5 这个地方取值就是0或者1

            all_predictions.extend(pred.cpu().numpy().flatten())#对应之前的：这一批次预测laebl加进去数组
            correct += pred.eq(target.data.view_as(pred)).cpu().sum()
            all_targets.extend(target.cpu().numpy().flatten())#对应之前的：这一批次实际label加进去数组

            # 添加内容方便计算 BSS BS
            all_predictions_y_true.extend(target.cpu().numpy().flatten())#对应之前的：拿到实际的label
            pos_prob = output  # 正类概率 [batch_size, 1]
            neg_prob = 1.0 - pos_prob  # 负类概率 [batch_size, 1]
            probabilities = torch.cat((neg_prob, pos_prob), dim=1)  # [batch_size, 2]，[负类概率, 正类概率]
            all_predictions_y_prob.extend(probabilities.cpu().numpy().tolist())  # 存储 [负类概率, 正类概率] 列表 #对应之前的： 拿到概率数值加入数组

            batch_count += 1

        metric = Metric(all_targets, all_predictions)
        print(metric.Matrix())
        TSS = metric.TSS()[0]
        print(f"TSS: {TSS}")
        test_loss /= batch_count
        accuracy = correct / len(data_x)
        message = (
            '\nTesting: Average loss: {:.4f}, Accuracy: {}/{} ({:.0f}%)\n'.format(
                test_loss, correct, len(data_x), 100. * accuracy))
        print("本轮评估完成", message)
        return {'loss': test_loss, 'accuracy': accuracy, 'tss': TSS, "metric": metric}, \
               all_predictions_y_true, all_predictions_y_prob


def evalual_all(data_x, data_y, model,batch_size):
    model.eval()
    test_loss = 0
    correct = 0
    all_targets = []
    all_predictions = []
    batch_count = 0
    all_predictions_y_true = []
    all_predictions_y_prob = []

    with torch.no_grad():
        for data, target in get_batches_all(data_x, data_y, batch_size):
            if not isinstance(data, torch.Tensor):
                data = torch.tensor(data, dtype=torch.float32)
            # 修改 target 处理：无论是否为 Tensor，都强制转换为 float32
            target = torch.tensor(target, dtype=torch.float32) if not isinstance(target, torch.Tensor) else target.to(dtype=torch.float32)

            if args.cuda:
                data, target = data.cuda(), target.cuda()
                class_weights_cuda = torch.tensor([class_weight[1.]], dtype=torch.float32).cuda()
            else:
                class_weights_cuda = torch.tensor([class_weight[1.]], dtype=torch.float32)

            output = model(data)  # [batch_size, 1]，概率值
            target = target.view(-1, 1)
            test_loss += F.binary_cross_entropy(output, target, weight=class_weights_cuda).item()

            pred = (output > 0.5).float()  # 阈值 0.5 这个地方取值就是0或者1

            all_predictions.extend(pred.cpu().numpy().flatten())  # 对应之前的：这一批次预测laebl加进去数组
            correct += pred.eq(target.data.view_as(pred)).cpu().sum()
            all_targets.extend(target.cpu().numpy().flatten())  # 对应之前的：这一批次实际label加进去数组

            # 添加内容方便计算 BSS BS
            all_predictions_y_true.extend(target.cpu().numpy().flatten())  # 对应之前的：拿到实际的label
            pos_prob = output  # 正类概率 [batch_size, 1]
            neg_prob = 1.0 - pos_prob  # 负类概率 [batch_size, 1]
            probabilities = torch.cat((neg_prob, pos_prob), dim=1)  # [batch_size, 2]，[负类概率, 正类概率]
            all_predictions_y_prob.extend(probabilities.cpu().numpy().tolist())  # 存储 [负类概率, 正类概率] 列表 #对应之前的： 拿到概率数值加入数组

            batch_count += 1

        metric = Metric(all_targets, all_predictions)
        print(metric.Matrix())
        TSS = metric.TSS()[0]
        print(f"TSS: {TSS}")
        test_loss /= batch_count
        accuracy = correct / len(data_x)
        message = (
            '\nTesting: Average loss: {:.4f}, Accuracy: {}/{} ({:.0f}%)\n'.format(
                test_loss, correct, len(data_x), 100. * accuracy))
        print("本轮评估完成", message)
        return {'loss': test_loss, 'accuracy': accuracy, 'tss': TSS, "metric": metric}, \
               all_predictions_y_true, all_predictions_y_prob


def read_parameters():
    # 初始化参数收集器
    parser = argparse.ArgumentParser(description='Time-LLM')
    # 数据集部分参数
    parser.add_argument('--num_dataset', type=int, default=9)  # 循环处理0-9个数据集
    parser.add_argument('--input_size', type=int, default=10)  # 循环处理0-9个数据集
    parser.add_argument('--time_step', type=int, default=40)  # 循环处理0-9个数据集
    parser.add_argument('--batch_size', type=int, default=16, help='batch size of train input data')
    parser.add_argument('--device', type=str, default="cuda")  # 循环处理0-9个数据集
    parser.add_argument('--datasetname', type=str, default="data", help='new_data_scaler,data')
    parser.add_argument('--epochs', type=int, default=50)
    # 公共训练参数
    parser.add_argument('--cuda', action='store_false')
    parser.add_argument('--lr', type=float, default=0.001)
    parser.add_argument('--optim', type=str, default='Adam')
    parser.add_argument('--seed', type=int, default=2021, help='random seed')
    parser.add_argument('--model_type', type=str, default='LLMFlareNet_1',
                        help='Onefitall,LLMFlareNet')
    parser.add_argument('--bert_emb', type=int, default=768) #不能改BERT-base:768
    parser.add_argument('--d_llm', type=int, default=768) #不能改BERT-base:768
    parser.add_argument('--d_model', type=int, default=16, help='patch of out_channels')
    # LLMFlareNetModel训练参数
    parser.add_argument('--bert_num_hidden_layers', type=int, default=2)
    parser.add_argument('--description_data', type=str,
                        default="数据形状是40*10,由40个耀斑物理特征时间步数据组成，每个时间步有10个特征,"
                                "每一组数据分别对应未来24小时内爆发的耀斑类别是否是大于等于M类别")
    parser.add_argument('--description_task', type=str,
                        default="使用这些数据预报未来24小时内爆发大于等于M类别耀斑的概率，"
                                "预报的概率值大于0.5则视为发生了")
    parser.add_argument('--n_heads', type=int, default=8, help='num of heads')
    parser.add_argument('--dropout', type=float, default=0.5, help='dimension of fcn')
    parser.add_argument('--num_tokens', type=int, default=1000, help='映射与时间有关的')
    parser.add_argument('--patch_len', type=int, default=1, help='patch length')#8
    parser.add_argument('--stride', type=int, default=1, help='stride')#5
    # OnefitallModel训练参数

    #NN输出层
    parser.add_argument('--batch_norm64_dim', type=int, default=64, help='Dimension for second batch norm layer')
    parser.add_argument('--batch_norm32_dim', type=int, default=32, help='Dimension for third batch norm layer')
    parser.add_argument('--dropout_rate', type=float, default=0.5, help='Dropout rate')
    parser.add_argument('--fc64_dim', type=int, default=64, help='Dimension for first fully connected layer')
    parser.add_argument('--fc32_dim', type=int, default=32, help='Dimension for second fully connected layer')
    parser.add_argument('--output_dim', type=int, default=1, help='Output dimension (number of classes)')
    # 备注参数
    parser.add_argument('--conmment', type=str, default="None")

    args = parser.parse_args()
    return args

def save_args_to_csv(args, model_base):
    """将args参数保存到CSV文件中，方便调参记录"""
    args_dict = vars(args)  # 将args转换为字典
    
    # 添加时间戳
    args_dict['timestamp'] = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
    args_dict['model_output_path'] = model_base
    
    # 创建DataFrame
    df = pd.DataFrame([args_dict])
    
    # 保存到model_base目录下
    csv_path = f"{model_base}/args_config.csv"
    df.to_csv(csv_path, index=False)
    print(f"参数配置已保存到: {csv_path}")
    
    # 同时保存到项目根目录的汇总文件
    summary_csv = "args_history.csv"
    if os.path.exists(summary_csv):
        # 如果文件存在，追加数据
        df.to_csv(summary_csv, mode='a', header=False, index=False)
    else:
        # 如果文件不存在，创建新文件
        df.to_csv(summary_csv, index=False)
    print(f"参数配置已追加到汇总文件: {summary_csv}")
    
    return csv_path


device = torch.device("cuda" if torch.cuda.is_available() else "cpu")



def get_model(model_name):
    # 在主程序中定义模型映射
    model_dict = {
        "LLMFlareNet_1": LLMFlareNet_1Model,
        "LLMFlareNet_2": LLMFlareNet_2Model,
        "Onefitall_11": Onefitall_11Model,
        "Onefitall_12": Onefitall_12Model,
        "Onefitall_13": Onefitall_13Model,
    }

    # 实例化模型
    model_class = model_dict.get(model_name)
    if model_class is None:
        raise ValueError(f"Unknown model type: {args.model_type}")
    return model_class

if __name__ == "__main__":

            args=read_parameters()

            # 当前训练备注
            commment =args.model_type

            start_time = time.time()

            timelabel = time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime(time.time()))

            # 定义训练输出目录
            model_base = f"./model_output/{timelabel}"
            os.makedirs(f"{model_base}")
            os.makedirs(f"{model_base}/plot")
            
            # 保存args参数到CSV
            save_args_to_csv(args, model_base)
            
            results_filepath = f"{model_base}/important.txt"
            results_logfilepath = f"{model_base}/log.txt"

            #保证输出到txt和控制台
            sys.stdout = DualOutput(results_logfilepath)

            # 定义十个数据集合存储内容
            all_matrix = np.array([[0, 0], [0, 0]])
            data_Recall, data_Precision, data_Accuracy, data_TSS, data_BSS, data_HSS, data_FAR = [], [], [], [], [], [], []

            # 打开文件准备写入
            with open(results_filepath, 'w') as results_file:

                # 在循环外部初始化存储损失的列表，每个列表包含所有数据集的损失数据，便于保存每一次损失
                all_epoch_losses = [[] for _ in range(args.num_dataset + 1)]
                all_val_losses = [[] for _ in range(args.num_dataset + 1)]

                for count in range(args.num_dataset + 1):  # 循环处理0-9个数据集
                    setup_seed_torch(args.seed)

                    train_csv_path = rf"./data/{count}Train.csv"
                    validate_csv_path = rf"./data/{count}Val.csv"
                    test_csv_path = rf"./data/{count}Test.csv"

                    train_x, train_y, validate_x, validate_y, test_x, test_y, class_weight = \
                        Preprocess(train_csv_path, validate_csv_path, test_csv_path)

                    model_filename = f"{model_base}/model_{count}.pt"  # 使用不同的文件名保存模型

                    # 保证每一个数据集的参数独立互不干扰，获取到lr便于后
                    # 面学习率步数衰减,提取其他参数方便其他使用
                    lr = args.lr
                    batch_size=args.batch_size

                    model=get_model(args.model_type)
                    model_train = model(args).float()
                    # model_train = model(args).to(torch.bfloat16)
                    model_train.to(device)


                    optimizer = getattr(optim, args.optim)(model_train.parameters(), lr=lr)

                    best_tss = -1
                    # 在训练开始之前，计算第0轮的训练集和验证集损失
                    print("==================== 第0轮评估训练集和验证集数据 ====================")
                    initial_train_metrics, _, _ = evalual_all(train_x, train_y, model_train,batch_size)  # 使用训练集评估当前模型
                    initial_val_metrics, _, _ = evalual_all(validate_x, validate_y, model_train,batch_size)  # 使用验证集评估当前模型

                    print(f"初始训练集损失: {initial_train_metrics['loss']}, 准确率: {initial_train_metrics['accuracy']}")
                    print(f"初始验证集损失: {initial_val_metrics['loss']}, 准确率: {initial_val_metrics['accuracy']}")
                    # 记录初始损失
                    all_epoch_losses[count].append(initial_train_metrics['loss'])
                    all_val_losses[count].append(initial_val_metrics['loss'])

                    # 初始化损失列表并将第0轮损失加入列表
                    epoch_losses = [initial_train_metrics['loss']]  # 第0轮训练集损失
                    val_losses = [initial_val_metrics['loss']]  # 第0轮验证集损失

                    for epoch in range(1, args.epochs + 1):

                        train_shuffle_x, train_shuffle_y = shuffle_data(train_x, train_y)
                        # 训练一个 epoch
                        train_loss = train_integer(epoch, train_shuffle_x, train_shuffle_y, optimizer, model_train, batch_size)
                        epoch_losses.append(train_loss)
                        all_epoch_losses[count].append(train_loss)# 记录这轮次到当前数据集里面

                        print(f"====================数据集{count}第{epoch}验证集轮评估数据=============================================")
                        val_metrics, _, _ = evalual_all(validate_x, validate_y, model_train,batch_size)  # 验证当前模型
                        val_losses.append(val_metrics['loss'])
                        all_val_losses[count].append(val_metrics['loss'])# 记录这轮次到当前数据集里面
                        print(f"====本轮验证集==========")
                        print(f"目前最佳TSS：{best_tss}")
                        print(f"本轮验证集TSS：{val_metrics['tss']}")

                        if val_metrics['tss'] > best_tss:
                            best_tss = val_metrics['tss']
                            save_torchModel(model_train, model_filename)  # 保存当前最佳模型

                        #  学习率衰减策略
                        if epoch % 10 == 0:
                            lr /= 10
                            for param_group in optimizer.param_groups:
                                param_group['lr'] = lr
                    #释放模型准备下一个数据集实例化
                    del model_train


                    print(f"====================数据集{count}测试集轮评估数据=============================================")
                    # 加载最佳模型
                    model_test = torch.load(model_filename)
                    # 测试集评估
                    test_metrics, all_predictions_y_true, all_predictions_y_prob =\
                        evalual_all(test_x, test_y,model_test,batch_size)
                    del model_test
                    # 清理 GPU 内存
                    # torch.cuda.empty_cache()

                    # 计算测试集矩阵
                    testMetrics = test_metrics["metric"]
                    print(testMetrics.Matrix())
                    all_matrix += testMetrics.Matrix()

                    data_Recall.append(testMetrics.Recall())
                    print("Recall", testMetrics.Recall())

                    data_Precision.append(testMetrics.Precision())
                    print("Precision", testMetrics.Precision())

                    data_Accuracy.append(testMetrics.Accuracy())
                    print("Accuracy", testMetrics.Accuracy())

                    data_TSS.append(testMetrics.TSS())
                    print("TSS", testMetrics.TSS())

                    data_HSS.append(testMetrics.HSS())
                    print("HSS", testMetrics.HSS())

                    data_FAR.append(testMetrics.FAR())
                    print("FAR", testMetrics.FAR())

                    # 开始求BSS
                    y_true = all_predictions_y_true
                    y_prob = np.array([row[1] for row in all_predictions_y_prob])

                    BS, BSS = BS_BSS_score(y_true, y_prob)
                    data_BSS.append([BS, BSS])
                    print("BS, BSS", [BS, BSS])

                    print(f"数据集 {count} 测试集TSS:", test_metrics['tss'])
                    print(f"数据集 {count} 验证集最优TSS:", best_tss)

                    # 写入结果到文件
                    results_file.write(f"数据集 {count} 测试集TSS: {test_metrics['tss']}\n")
                    results_file.write(f"数据集 {count} 验证集最优TSS: {best_tss}\n")
                    results_file.write("=================================================================\n")

                    # 绘制并保存损失曲线
                    plot_losses(epoch_losses, count, 'train', fr"{args.model_type}", path=f"{model_base}/plot")
                    plot_losses(val_losses, count, 'validation', rf"{args.model_type}", path=f"{model_base}/plot")

                # 在循环结束后将所有损失数据写入CSV文件
                df_train_losses = pd.DataFrame(all_epoch_losses).T  # 转置以使每列代表一个数据集的损失
                df_val_losses = pd.DataFrame(all_val_losses).T

                # 保存到CSV文件
                df_train_losses.to_csv(fr"{model_base}/{args.model_type}_train_loss.csv", index=False)
                df_val_losses.to_csv(fr"{model_base}/{args.model_type}_validation_loss.csv", index=False)

                print("#接下来计算所有测试集指标均值和方法")
                print(all_matrix)
                print(data_BSS)
                # 转换数据为numpy数组以便计算
                data_Recall = np.array(data_Recall)
                data_Precision = np.array(data_Precision)
                data_Accuracy = np.array(data_Accuracy)
                data_TSS = np.array(data_TSS)
                data_HSS = np.array(data_HSS)
                data_FAR = np.array(data_FAR)
                data_BSS = np.array(data_BSS)
                # 计算均值和标准差
                results = {
                    "Metric": ["Recall", "Precision", "Accuracy", "TSS", "HSS", "FAR", "BSS"],
                    "Mean": [data_Recall.mean(axis=0), data_Precision.mean(axis=0), data_Accuracy.mean(axis=0),
                             data_TSS.mean(axis=0), data_HSS.mean(axis=0), data_FAR.mean(axis=0), data_BSS.mean(axis=0)],
                    "Std": [data_Recall.std(axis=0), data_Precision.std(axis=0), data_Accuracy.std(axis=0),
                            data_TSS.std(axis=0), data_HSS.std(axis=0), data_FAR.std(axis=0), data_BSS.std(axis=0)]
                }

                # 将结果写入Excel
                df = pd.DataFrame(results)
                excel_filename = f'{model_base}/results.xlsx'
                df.to_excel(excel_filename, index=False)
                print(f"结果已写入 {excel_filename}")

                # 记录当前TSS
                # 设置汇总result.xlsx
                filename = 'result.xlsx'
                new_row = [f'{model_base}-{commment}', data_TSS.mean(axis=0)[1]]

                # 判断文件是否存在
                if os.path.exists(filename):
                    # 文件存在，加载文件
                    wb = openpyxl.load_workbook(filename)
                    ws = wb.active
                else:
                    # 文件不存在，创建一个新工作簿
                    wb = openpyxl.Workbook()
                    ws = wb.active

                # 获取当前行数
                current_row = ws.max_row + 1

                # 将数据写入新行
                for col_num, value in enumerate(new_row, start=1):
                    ws.cell(row=current_row, column=col_num, value=value)

                # 保存文件
                wb.save(filename)

            # 删除模型
            for filename in os.listdir(model_base):
                # 检查文件是否以 .pt 结尾
                if filename.endswith('.pt'):
                    # 构造文件的完整路径
                    file_path = os.path.join(model_base, filename)
                    try:
                        # 删除文件
                        os.remove(file_path)
                        pass
                    except Exception as e:
                        print(f"删除文件 {file_path} 时出错: {e}")

            end_time = time.time()
            elapsed_time_minutes = (end_time - start_time) / 60
            print(f"程序运行时间: {elapsed_time_minutes:.2f} 分钟")

